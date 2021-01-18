from http.server import BaseHTTPRequestHandler
import json
import re
import datetime
import openpyxl
from io import BytesIO
import urllib.request

# Source https://www.destatis.de/DE/Themen/Gesellschaft-Umwelt/Bevoelkerung/Bevoelkerungsstand/Tabellen/bevoelkerung-nichtdeutsch-laender.html
totalGermany = 83166711
states = {
  'Baden-Württemberg': {
    'total': 11100394
  },
  'Bayern': {
    'total': 13124737
  },
  'Berlin': {
    'total': 3669491
  },
  'Brandenburg': {
    'total': 2521893
  },
  'Bremen': {
    'total': 681202
  },
  'Hamburg': {
    'total': 1847253
  },
  'Hessen': {
    'total': 6288080
  },
  'Mecklenburg-Vorpommern': {
    'total': 1608138
  },
  'Niedersachsen': {
    'total': 7993608
  },
  'Nordrhein-Westfalen': {
    'total': 17947221
  },
  'Rheinland-Pfalz': {
    'total': 4093903
  },
  'Saarland': {
    'total':  986887
  },
  'Sachsen': {
    'total': 4071971
  },
  'Sachsen-Anhalt': {
    'total': 2194782
  },
  'Schleswig-Holstein': {
    'total': 2903773
  },
  'Thüringen': {
    'total': 2133378
  }
}

# Request to load excel sheet
url = 'https://www.rki.de/DE/Content/InfAZ/N/Neuartiges_Coronavirus/Daten/Impfquotenmonitoring.xlsx?__blob=publicationFile'
hdr = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11'}

req = urllib.request.Request(url, headers=hdr)
response = urllib.request.urlopen(req)
file = response.read()


# Read excel sheet
wb = openpyxl.load_workbook(filename = BytesIO(file)) 
sheet = wb[wb.sheetnames[1]]

# Load update time
lastUpdateRawString = wb.sheetnames[1]
relastUpdateMatch = re.search(r"[\d]{2}\.[\d]{2}\.[\d]{2}", lastUpdateRawString)
lastUpdate = datetime.datetime.strptime(relastUpdateMatch.group(), '%d.%m.%y')

# Load data from rows
sumStates = 0
sumDiffStates = 0
for row in sheet.iter_rows(max_row=19):
  if row[1].value is None:
    continue
  state = row[1].value.replace("*", "")
  if state in states:
    states[state]['rs'] = str(row[0].value)

    # First vaccination
    states[state]['vaccinated'] = row[2].value
    states[state]['by_accine'] = {}
    states[state]['by_accine']['biontech'] = row[3].value
    states[state]['by_accine']['moderna'] = row[4].value
    states[state]['difference_to_the_previous_day'] = row[5].value
    states[state]['vaccinations_per_1000_inhabitants'] = round(states[state]['vaccinated'] / states[state]['total'] * 1000, 2)
    states[state]['quote'] = row[6].value

    # Second vaccination
    states[state]['2nd-vaccination'] = {}
    states[state]['2nd-vaccination']['vaccinated'] = row[7].value
    states[state]['2nd-vaccination']['difference_to_the_previous_day'] = row[8].value

    sumStates += states[state]['vaccinated']
    sumDiffStates += states[state]['difference_to_the_previous_day']

res = {
  'lastUpdate': lastUpdate.isoformat(),
  'states': states,
  'vaccinated': sumStates,
  'difference_to_the_previous_day': sumDiffStates,
  'vaccinations_per_1000_inhabitants': round(sumStates / totalGermany * 1000, 2),
  'total': totalGermany,
  'quote': round(sumStates / totalGermany * 100, 2)
}

# HTTP handler
class handler(BaseHTTPRequestHandler):
  def do_GET(self):
    self.send_response(200)
    self.send_header('Content-Type', 'application/json')
    self.end_headers()
    self.wfile.write(json.dumps(res).encode())
    return
