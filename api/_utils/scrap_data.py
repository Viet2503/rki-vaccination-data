""" Scrap Data from Excel sheet """
import re
import datetime
from io import BytesIO
import urllib.request
import openpyxl
from .statics import inhabitants


def get_file():
  """ Get remote file content """
  # Request to load excel sheet
  url = 'https://www.rki.de/DE/Content/InfAZ/N/Neuartiges_Coronavirus/Daten' \
      '/Impfquotenmonitoring.xlsx?__blob=publicationFile'
  hdr = {
      'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11'
      ' (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11'}

  req = urllib.request.Request(url, headers=hdr)
  with urllib.request.urlopen(req) as response:
    return response.read()


def get_data():
  """ Get Data for API """
  states = inhabitants.STATES

  file = get_file()

  # Read excel sheet
  work_book = openpyxl.load_workbook(filename=BytesIO(file))
  sheet = work_book[work_book.sheetnames[2]]

  # Load update time
  first_sheet = work_book[work_book.sheetnames[0]]
  relast_update_match = re.search(
      r"[\d]{2}\.[\d]{2}\.[\d]{2}", first_sheet['A3'].value)
  last_update = datetime.datetime.strptime(
      relast_update_match.group(), '%d.%m.%y')

  # Load data from rows
  sum_states = 0
  sum_states2nd = 0
  sum_diff_states = 0
  sum_diff_states2nd = 0
  for row in sheet.iter_rows(max_row=21):
    if row[1].value is None:
      continue
    state = row[1].value.replace("*", "").strip()

    if state in states:
      states[state]['rs'] = str(row[0].value)

      # First vaccination
      states[state]['vaccinated'] = row[2].value
      states[state]['vaccinated_by_accine'] = {}
      states[state]['vaccinated_by_accine']['biontech'] = row[3].value
      states[state]['vaccinated_by_accine']['moderna'] = row[4].value
      states[state]['vaccinated_by_accine']['astrazeneca'] = row[5].value
      # states[state]['vaccinated_by_accine']['janssen'] = row[6].value
      states[state]['difference_to_the_previous_day'] = row[7].value
      states[state]['vaccinations_per_1000_inhabitants'] = round(states[state]['vaccinated']
                                                                / states[state]['total'] * 1000, 2)
      states[state]['quote'] = round(
        states[state]['vaccinated'] / states[state]['total'] * 100, 2)

      sum_states += states[state]['vaccinated']
      sum_diff_states += states[state]['difference_to_the_previous_day']

      # Second vaccination
      states[state]['2nd_vaccination'] = {}
      states[state]['2nd_vaccination']['vaccinated'] = row[8].value
      states[state]['2nd_vaccination']['vaccinated_by_accine'] = {}
      states[state]['2nd_vaccination']['vaccinated_by_accine']['biontech'] = row[9].value
      states[state]['2nd_vaccination']['vaccinated_by_accine']['moderna'] = row[10].value
      states[state]['2nd_vaccination']['vaccinated_by_accine']['astrazeneca'] = row[11].value
      states[state]['2nd_vaccination']['vaccinated_by_accine']['janssen'] = row[12].value
      states[state]['2nd_vaccination']['difference_to_the_previous_day'] = row[13].value
      states[state]['2nd_vaccination']['quote'] = round(
        states[state]['2nd_vaccination']['vaccinated'] / states[state]['total'] * 100, 2)

      sum_states2nd += states[state]['2nd_vaccination']['vaccinated']

      if states[state]['2nd_vaccination']['difference_to_the_previous_day'] is not None:
        sum_diff_states2nd += states[state]['2nd_vaccination']['difference_to_the_previous_day']

    elif state == 'Bundesressorts':
      sum_states += row[2].value
      sum_diff_states += row[7].value
      sum_states2nd += row[8].value
      sum_diff_states2nd += row[13].value

  return {
    "lastUpdate": last_update,
    "states": states,
    "sumStates": sum_states,
    "sumStates2nd": sum_states2nd,
    "sumDiffStates": sum_diff_states,
    "sumDiffStates2nd": sum_diff_states2nd,
    "totalGermany": inhabitants.TOTAL
  }
